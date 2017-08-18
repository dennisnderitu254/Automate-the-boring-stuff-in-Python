#! python3

# Chapter 16 Project Sending Member Dues Reminder Emails
# Sends emails based on payment status in a spreadsheet.

import smtplib
import sys
import openpyxl

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

smtpObj = smtplib.SMTP_SSL('smtp.gmail.com', 465)
smtpObj.ehlo()
smtpObj.login('myemail@address.com', sys.argv[1])

for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\n\nDear %s,\nRecords indicate that you have not paid dues for %s. Please pay ASAP! Thanks!" % (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('myemail@address.com', email, body)
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
smtpObj.quit()
